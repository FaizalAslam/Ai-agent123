import json
import sys
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

VENV_SITE = Path(__file__).resolve().parent.parent / ".venv" / "Lib" / "site-packages"
if VENV_SITE.exists():
    # The bundled test Python has the compiled Office libraries we need, while
    # the project venv has Flask. Preload Flask, then remove the venv path so
    # lxml/Pillow come from the bundled runtime instead of stale venv wheels.
    sys.path.insert(0, str(VENV_SITE))
    import flask  # noqa: F401
    import flask.testing as flask_testing
    import dotenv  # noqa: F401
    sys.path.remove(str(VENV_SITE))
    flask_testing._get_werkzeug_version = lambda: "3.x-test"

import server
from openpyxl import Workbook, load_workbook

from parser.command_complexity import classify_office_command_complexity
from utils.file_paths import extract_office_filename_hint, resolve_existing_office_path
from utils.office_actions import OfficeActionError, validate_actions


class OfficeOverhaulTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.command_map_path = Path("command_map.json")
        cls.original_command_map = cls.command_map_path.read_bytes() if cls.command_map_path.exists() else None
        server.system_core.open_path = lambda *args, **kwargs: True
        server.ui.manual_selector = lambda: ""
        cls.client = server.app.test_client()

    @classmethod
    def tearDownClass(cls):
        if cls.original_command_map is None:
            if cls.command_map_path.exists():
                cls.command_map_path.unlink()
        else:
            cls.command_map_path.write_bytes(cls.original_command_map)

    def payload(self, response):
        data = response.get_json(silent=True)
        self.assertIsInstance(data, dict)
        self.assertIn("success", data)
        self.assertIn("status", data)
        self.assertIn("intent", data)
        return data

    def test_office_execute_route_protects_from_manual_selector(self):
        data = self.payload(self.client.post("/execute", json={"command": "create a new Excel file"}))
        self.assertTrue(data["success"], data)
        self.assertEqual(data["intent"], "office_automation")
        self.assertFalse(data.get("requires_manual_selection"), data)
        path = Path(data["file_path"])
        self.assertTrue(path.exists(), data)
        self.assertEqual(path.suffix.lower(), ".xlsx")

    def test_complexity_classifier(self):
        self.assertEqual(classify_office_command_complexity("create a new Excel file"), "simple")
        self.assertEqual(classify_office_command_complexity("create workbook then bold header"), "compound_explicit")
        self.assertEqual(classify_office_command_complexity("create a professional spreadsheet with sample data"), "semantic_complex")

    def test_partial_planner_falls_through_to_openai_then_last_resort(self):
        with patch("server.parse_command", return_value=[]):
            _cache_key, actions, source, error, plan = server._resolve_actions(
                "excel",
                "create workbook; do the impossible custom chart thing",
            )
        self.assertIsNone(error)
        self.assertEqual(source, "planner-partial")
        self.assertTrue(actions)
        self.assertTrue(plan["diag"]["openai_attempted"], plan)

    def test_parser_validation_failure_falls_through_to_openai(self):
        fake_plan = SimpleNamespace(
            actions=[],
            requires_api=True,
            success=False,
            errors=["planner miss"],
            clauses=[],
            to_dict=lambda: {
                "success": False,
                "requires_api": True,
                "actions": [],
                "errors": ["planner miss"],
                "clauses": [],
            },
        )
        fake_ai = SimpleNamespace(
            success=True,
            actions=[{"action": "create_workbook"}],
            error_code="",
            message="ok",
            raw_response_preview="",
            output_filename="",
            ai_context={},
        )
        with patch("server.command_map.get_cached_actions", return_value=(None, None, 0)), \
             patch("server.command_map.save_actions", return_value=True), \
             patch("server.plan_office_command", return_value=fake_plan), \
             patch("server.parse_command", return_value=[{"action": "set_bg_color", "range": "WIDTHS:WIDTHS", "color": "yellow"}]), \
             patch.object(server._openai_handler, "interpret_result", return_value=fake_ai):
            _cache_key, actions, source, error, plan = server._resolve_actions("excel", "bad parser range")
        self.assertIsNone(error)
        self.assertEqual(source, "openai-fallback")
        self.assertEqual(actions, [{"action": "create_workbook"}])
        self.assertIn("parser:", " | ".join(plan["diag"]["validation_errors"]))

    def test_excel_headers_override_defaults(self):
        data = self.payload(self.client.post(
            "/office/execute",
            json={"app": "excel", "raw": "create workbook; add table with headers Name, Amount, Status and 5 rows"},
        ))
        self.assertTrue(data["success"], data)
        ws = load_workbook(data["file_path"]).active
        self.assertEqual([ws["A1"].value, ws["B1"].value, ws["C1"].value], ["Name", "Amount", "Status"])

    def test_validator_rejects_pseudo_ranges_placeholders_and_bad_colors(self):
        known = server._known_office_actions("excel")
        with self.assertRaises(OfficeActionError) as bad_range:
            validate_actions("excel", [{"action": "set_bg_color", "range": "WIDTHS:WIDTHS", "color": "yellow"}], known_actions=known)
        self.assertEqual(bad_range.exception.error_code, "INVALID_EXCEL_RANGE")
        with self.assertRaises(OfficeActionError) as bad_placeholder:
            validate_actions("excel", [{"action": "write_cell", "cell": "A1", "value": "{unresolved}"}], known_actions=known)
        self.assertEqual(bad_placeholder.exception.error_code, "INVALID_ACTION")
        with self.assertRaises(OfficeActionError) as bad_color:
            validate_actions("excel", [{"action": "set_bg_color", "range": "A1", "color": "not-a-color"}], known_actions=known)
        self.assertEqual(bad_color.exception.error_code, "INVALID_COLOR")

    def test_open_filename_hint_and_case_insensitive_resolution(self):
        folder = Path("outputs/office/excel").resolve()
        folder.mkdir(parents=True, exist_ok=True)
        target = folder / "Case_File.xlsx"
        wb = Workbook()
        wb.active["A1"] = "ok"
        wb.save(target)

        self.assertEqual(
            extract_office_filename_hint("open the excel file named Case_File on desktop", "excel"),
            "Case_File",
        )
        resolved = resolve_existing_office_path(
            "case_file",
            "excel",
            base_dir=folder,
            command_text="open workbook case_file",
        )
        self.assertEqual(resolved, target)

    def test_office_response_contract_contains_parser_and_results(self):
        data = self.payload(self.client.post("/office/execute", json={"app": "word", "raw": "create a Word document"}))
        self.assertTrue(data["success"], data)
        self.assertEqual(data["intent"], "office_automation")
        self.assertIn("parser_used", data)
        self.assertIn("results", data)
        self.assertIn("data", data)
        self.assertEqual(data["data"]["parser_used"], data["parser_used"])


if __name__ == "__main__":
    unittest.main()
