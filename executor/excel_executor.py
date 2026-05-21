# modules/excel_executor.py
import logging
import re

logger = logging.getLogger("OfficeAgent")


def _normalize_excel_argb(color):
    color = str(color or "").strip().lstrip("#").upper()
    named_colors = {
        "RED": "FFFF0000",
        "GREEN": "FF00B050",
        "BLUE": "FF0070C0",
        "YELLOW": "FFFFFF00",
        "ORANGE": "FFFFA500",
        "PURPLE": "FF7030A0",
        "PINK": "FFFF69B4",
        "BLACK": "FF000000",
        "WHITE": "FFFFFFFF",
        "GRAY": "FF808080",
        "GREY": "FF808080",
        "DARK RED": "FFC00000",
        "DARK BLUE": "FF00008B",
        "DARK GREEN": "FF006400",
        "LIGHT BLUE": "FFADD8E6",
        "LIGHT GRAY": "FFD3D3D3",
        "TEAL": "FF008080",
        "CYAN": "FF00FFFF",
        "MAGENTA": "FFFF00FF",
        "GOLD": "FFFFD700",
        "BROWN": "FFA52A2A",
        "NAVY": "FF000080",
    }
    if color in named_colors:
        return named_colors[color]
    if re.fullmatch(r"[0-9A-F]{6}", color):
        return "FF" + color
    if re.fullmatch(r"[0-9A-F]{8}", color):
        return color
    raise ValueError(f"Invalid Excel color: {color}")


def _xl_color(hex_color):
    from openpyxl.styles import Color
    return Color(rgb=_normalize_excel_argb(hex_color))


class ExcelExecutor:
    def __init__(self, wb, ws):
        self.wb = wb
        self.ws = ws

    def _sheet_for_action(self, p):
        sheet_name = str(p.get("sheet_name") or p.get("name") or "").strip()
        if not sheet_name:
            return self.ws
        if sheet_name not in self.wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")
        return self.wb[sheet_name]

    def _set_current_sheet(self, ws):
        self.ws = ws
        self.wb.active = self.wb.worksheets.index(ws)

    def _iter_cells(self, range_ref):
        ref = (range_ref or "").strip()
        if not ref:
            return []
        if ":" not in ref:
            return [self.ws[ref]]
        cells = []
        for row in self.ws[ref]:
            for cell in row:
                cells.append(cell)
        return cells

    def run(self, action_dict):
        action  = action_dict.get("action", "unknown")
        handler = getattr(self, f"_do_{action}", None)
        if not handler:
            logger.warning(f"Excel: Unknown action '{action}'")
            return {"status": "failed", "action": action, "message": f"Unknown action: {action}", "error_code": "UNKNOWN_ACTION"}
        try:
            handler(action_dict)
            return {"status": "success", "action": action, "message": ""}
        except Exception as e:
            logger.error(f"Excel action '{action}' failed: {e}")
            return {"status": "failed", "action": action, "message": str(e), "error_code": "ACTION_EXECUTION_ERROR"}

    # â”€â”€ Cell / Range Operations â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    def _do_write_cell(self, p):
        self.ws[p["cell"]] = p.get("value", "")

    def _do_write_formula(self, p):
        self.ws[p["cell"]] = p.get("formula", "")

    def _do_write_range(self, p):
        import openpyxl
        start = p.get("start_cell", "A1")
        values = p.get("values", [])
        try:
            start_row, start_col = openpyxl.utils.cell.coordinate_to_tuple(start)
        except Exception:
            logger.warning("_do_write_range: invalid start_cell %r", start)
            return False
        if isinstance(values, list):
            for i, row_data in enumerate(values):
                if isinstance(row_data, list):
                    for j, val in enumerate(row_data):
                        self.ws.cell(row=start_row + i, column=start_col + j).value = val
                else:
                    self.ws.cell(row=start_row + i, column=start_col).value = row_data

    def _do_read_cell(self, p):
        val = self.ws[p["cell"]].value
        logger.info(f"Cell {p['cell']} = {val}")
        return val

    def _do_clear_cell(self, p):
        self.ws[p["cell"]].value = None

    def _do_clear_range(self, p):
        for row in self.ws[p["range"]]:
            for cell in row:
                cell.value = None

    def _do_clear_all(self, p):
        for row in self.ws.iter_rows():
            for cell in row:
                cell.value = None

    def _do_clear_format(self, p):
        from openpyxl.styles import Font, PatternFill, Alignment, Border
        for row in self.ws[p["range"]]:
            for cell in row:
                cell.font      = Font()
                cell.fill      = PatternFill()
                cell.alignment = Alignment()
                cell.border    = Border()

    def _do_copy_range(self, p):
        logger.info(f"Copy range {p.get('range')} (clipboard â€” requires win32com)")

    def _do_cut_range(self, p):
        logger.info(f"Cut range {p.get('range')} (clipboard â€” requires win32com)")

    def _do_paste_range(self, p):
        logger.info(f"Paste to {p.get('cell')} (clipboard â€” requires win32com)")

    def _do_paste_values_only(self, p):
        logger.info(f"Paste values only to {p.get('cell')} (requires win32com)")

    def _do_undo(self, p):
        logger.info("Undo (requires win32com)")

    def _do_redo(self, p):
        logger.info("Redo (requires win32com)")

    # â”€â”€ Font & Style â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    def _do_set_bold(self, p):
        from openpyxl.styles import Font
        for cell in self._iter_cells(p["range"]):
            cell.font = Font(
                bold=p.get("bold", True),
                italic=cell.font.italic,
                size=cell.font.size,
                name=cell.font.name,
                color=cell.font.color
            )

    def _do_set_italic(self, p):
        from openpyxl.styles import Font
        for cell in self._iter_cells(p["range"]):
            cell.font = Font(
                italic=p.get("italic", True),
                bold=cell.font.bold,
                size=cell.font.size,
                name=cell.font.name,
                color=cell.font.color,
            )

    def _do_set_underline(self, p):
        from openpyxl.styles import Font
        val = "single" if p.get("underline", True) else None
        for row in self.ws[p["range"]]:
            for cell in row:
                cell.font = Font(
                    underline=val,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    size=cell.font.size,
                    color=cell.font.color,
                )

    def _do_set_strikethrough(self, p):
        from openpyxl.styles import Font
        for row in self.ws[p["range"]]:
            for cell in row:
                cell.font = Font(
                    strike=True,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    size=cell.font.size,
                    color=cell.font.color,
                )

    def _do_set_font_size(self, p):
        from openpyxl.styles import Font
        for cell in self._iter_cells(p["range"]):
            cell.font = Font(
                size=int(p["size"]),
                bold=cell.font.bold,
                italic=cell.font.italic,
                name=cell.font.name
            )

    def _do_set_font_name(self, p):
        from openpyxl.styles import Font
        for cell in self._iter_cells(p["range"]):
            cell.font = Font(
                name=p["name"],
                bold=cell.font.bold,
                italic=cell.font.italic,
                size=cell.font.size
            )

    def _do_set_font_color(self, p):
        from openpyxl.styles import Font
        color = _xl_color(p["color"])
        for cell in self._iter_cells(p["range"]):
            cell.font = Font(
                color=color,
                bold=cell.font.bold,
                italic=cell.font.italic,
                size=cell.font.size,
                name=cell.font.name
            )

    def _do_set_bg_color(self, p):
        from openpyxl.styles import PatternFill
        color = _normalize_excel_argb(p["color"])
        fill = PatternFill(
            start_color=color,
            end_color=color,
            fill_type="solid"
        )
        for cell in self._iter_cells(p["range"]):
            cell.fill = fill

    def _do_set_border(self, p):
        from openpyxl.styles import Border, Side
        side   = Side(style=p.get("style", "thin"))
        border = Border(left=side, right=side, top=side, bottom=side)
        for row in self.ws[p["range"]]:
            for cell in row:
                cell.border = border

    def _do_remove_border(self, p):
        from openpyxl.styles import Border
        for row in self.ws[p["range"]]:
            for cell in row:
                cell.border = Border()

    def _do_set_alignment(self, p):
        from openpyxl.styles import Alignment
        horiz = p.get("alignment", "left")
        for row in self.ws[p["range"]]:
            for cell in row:
                cell.alignment = Alignment(
                    horizontal=horiz,
                    vertical=cell.alignment.vertical,
                    wrap_text=cell.alignment.wrap_text
                )

    def _do_set_vertical_alignment(self, p):
        from openpyxl.styles import Alignment
        for row in self.ws[p["range"]]:
            for cell in row:
                cell.alignment = Alignment(
                    vertical=p.get("alignment", "center"),
                    horizontal=cell.alignment.horizontal
                )

    def _do_set_wrap_text(self, p):
        from openpyxl.styles import Alignment
        for row in self.ws[p["range"]]:
            for cell in row:
                cell.alignment = Alignment(
                    wrap_text=p.get("wrap", True),
                    horizontal=cell.alignment.horizontal
                )

    def _do_set_number_format(self, p):
        for cell in self._iter_cells(p["range"]):
            cell.number_format = p.get("format", "General")

    # â”€â”€ Merge / Unmerge â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    def _do_merge_cells(self, p):
        self.ws.merge_cells(p["range"])

    def _do_unmerge_cells(self, p):
        self.ws.unmerge_cells(p["range"])

    # â”€â”€ Rows / Columns â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    def _do_insert_row(self, p):
        row = int(p["row"])
        amount = max(1, int(p.get("amount") or p.get("count") or 1))
        self.ws.insert_rows(row, amount)
        self._shift_tables_for_insert_rows(row, amount)

    def _do_insert_column(self, p):
        from openpyxl.utils import column_index_from_string
        col = column_index_from_string(str(p["column"]).upper())
        amount = max(1, int(p.get("amount") or p.get("count") or 1))
        self.ws.insert_cols(col, amount)
        self._shift_tables_for_insert_cols(col, amount)

    def _do_delete_row(self, p):
        self.ws.delete_rows(int(p["row"]))

    def _do_delete_column(self, p):
        from openpyxl.utils import column_index_from_string
        self.ws.delete_cols(column_index_from_string(str(p["column"]).upper()))

    def _do_set_row_height(self, p):
        self.ws.row_dimensions[int(p["row"])].height = float(p["height"])

    def _do_set_column_width(self, p):
        self.ws.column_dimensions[p["column"].upper()].width = float(p["width"])

    def _do_autofit_columns(self, p):
        target_range = str(p.get("range") or p.get("columns") or "").upper().strip()
        if target_range:
            # Determine column letter bounds from range like "A:C" or "A1:C5"
            col_match = re.match(r"^([A-Z]{1,3})(?:\d*)?:([A-Z]{1,3})(?:\d*)?$", target_range)
            if col_match:
                from openpyxl.utils import column_index_from_string, get_column_letter
                start_idx = column_index_from_string(col_match.group(1))
                end_idx   = column_index_from_string(col_match.group(2))
                col_letters = [get_column_letter(i) for i in range(start_idx, end_idx + 1)]
            else:
                col_letters = [col[0].column_letter for col in self.ws.columns]
        else:
            col_letters = [col[0].column_letter for col in self.ws.columns]

        for col in self.ws.columns:
            letter = col[0].column_letter
            if letter not in col_letters:
                continue
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            self.ws.column_dimensions[letter].width = min(max_len + 4, 60)

    def _do_autofit_rows(self, p):
        logger.info("Autofit rows (approximate â€” full support requires win32com)")
        for row in self.ws.iter_rows():
            self.ws.row_dimensions[row[0].row].height = 15

    def _do_hide_row(self, p):
        self.ws.row_dimensions[int(p["row"])].hidden = True

    def _do_unhide_row(self, p):
        self.ws.row_dimensions[int(p["row"])].hidden = False

    def _do_hide_column(self, p):
        self.ws.column_dimensions[p["column"].upper()].hidden = True

    def _do_unhide_column(self, p):
        self.ws.column_dimensions[p["column"].upper()].hidden = False

    # â”€â”€ Sheet Operations â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    def _do_add_sheet(self, p):
        name = str(p.get("name") or "Sheet").strip() or "Sheet"
        ws = self.wb[name] if name in self.wb.sheetnames else self.wb.create_sheet(title=name)
        # Treat "add worksheet named X" as selecting X for following actions.
        self._set_current_sheet(ws)

    def _do_delete_sheet(self, p):
        name = p.get("name")
        if name in self.wb.sheetnames:
            del self.wb[name]

    def _do_rename_sheet(self, p):
        old_name = p.get("old_name")
        new_name = p.get("new_name", "Sheet")
        if old_name and old_name in self.wb.sheetnames:
            ws = self.wb[old_name]
            ws.title = new_name
            self._set_current_sheet(ws)
            return
        # If old name is not provided, rename the active sheet.
        self.ws.title = new_name

    def _do_duplicate_sheet(self, p):
        src = self.wb[p.get("name", self.ws.title)]
        self.wb.copy_worksheet(src)

    def _do_hide_sheet(self, p):
        self.wb[p["name"]].sheet_state = "hidden"

    def _do_unhide_sheet(self, p):
        self.wb[p["name"]].sheet_state = "visible"

    def _do_set_active_sheet(self, p):
        self._set_current_sheet(self.wb[p["name"]])

    def _do_move_sheet(self, p):
        self.wb.move_sheet(p.get("name"), offset=int(p.get("position", 0)))

    def _do_protect_sheet(self, p):
        ws = self._sheet_for_action(p)
        password = p.get("password", "")
        ws.protection.sheet = True
        if password:
            ws.protection.set_password(str(password))
        logger.info("Protected sheet '%s'", ws.title)

    def _do_unprotect_sheet(self, p):
        from openpyxl.worksheet.protection import SheetProtection
        ws = self._sheet_for_action(p)
        ws.protection = SheetProtection(sheet=False)
        logger.info("Unprotected sheet '%s'", ws.title)

    def _do_protect_workbook(self, p):
        password = str(p.get("password") or "")
        self.wb.security.lockStructure = True
        if password and hasattr(self.wb.security, "set_workbook_password"):
            self.wb.security.set_workbook_password(password)
        logger.info("Workbook structure protected")

    def _do_unprotect_workbook(self, p):
        self.wb.security.lockStructure = False
        self.wb.security.workbookPassword = ""
        logger.info("Workbook structure unprotected")

    # â”€â”€ Data Tools â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    def _do_freeze_panes(self, p):
        self.ws.freeze_panes = p["cell"]

    def _do_unfreeze_panes(self, p):
        self.ws.freeze_panes = None

    def _do_sort_range(self, p):
        reverse = p.get("order", "ascending") == "descending"
        region  = list(self.ws[p["range"]])
        # Capture all values before sorting so we can write back in sorted order.
        snapshot = [[cell.value for cell in row] for row in region]
        snapshot.sort(key=lambda r: (r[0] is None, r[0] or ""), reverse=reverse)
        for row_cells, sorted_values in zip(region, snapshot):
            for cell, value in zip(row_cells, sorted_values):
                cell.value = value

    def _do_filter_range(self, p):
        self.ws.auto_filter.ref = p["range"]

    def _do_remove_filter(self, p):
        self.ws.auto_filter.ref = None

    def _do_find_replace(self, p):
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value and str(cell.value) == p.get("find_text", ""):
                    cell.value = p.get("replace_text", "")

    def _do_remove_duplicates(self, p):
        seen, rows_to_delete = set(), []
        for row in self.ws.iter_rows():
            key = tuple(c.value for c in row)
            if key in seen:
                rows_to_delete.append(row[0].row)
            else:
                seen.add(key)
        for row_idx in reversed(rows_to_delete):
            self.ws.delete_rows(row_idx)

    def _do_text_to_columns(self, p):
        logger.info(f"Text to columns with delimiter '{p.get('delimiter')}' (requires win32com)")

    def _do_create_named_range(self, p):
        from openpyxl.workbook.defined_name import DefinedName
        attr_text = f"'{self.ws.title}'!{p['range']}"
        self.wb.defined_names[p["name"]] = DefinedName(name=p["name"], attr_text=attr_text)

    def _do_add_conditional_formatting(self, p):
        from openpyxl.formatting.rule import ColorScaleRule
        rule = ColorScaleRule(
            start_type="min", start_color=_normalize_excel_argb("FF0000"),
            end_type="max",   end_color=_normalize_excel_argb("00FF00")
        )
        self.ws.conditional_formatting.add(p["range"], rule)

    def _do_add_data_validation(self, p):
        from openpyxl.worksheet.datavalidation import DataValidation
        vals = p.get("values", [])
        if isinstance(vals, list):
            vals = ",".join(str(v) for v in vals)
        dv = DataValidation(type="list", formula1=f'"{vals}"')
        self.ws.add_data_validation(dv)
        dv.add(self.ws[p["range"]])

    def _do_insert_comment(self, p):
        from openpyxl.comments import Comment
        self.ws[p["cell"]].comment = Comment(p.get("text", ""), "Agent")

    def _do_delete_comment(self, p):
        self.ws[p["cell"]].comment = None

    def _do_insert_hyperlink(self, p):
        cell           = self.ws[p["cell"]]
        cell.value     = p.get("text", p["cell"])
        cell.hyperlink = p.get("url", "")

    def _do_create_table(self, p):
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries

        ws = self._sheet_for_action(p)
        start = p.get("start_cell", "A1")
        start_row, start_col = coordinate_to_tuple(start)

        explicit_headers = p.get("headers") or []
        if isinstance(explicit_headers, str):
            explicit_headers = [h.strip() for h in explicit_headers.split(",") if h.strip()]

        used_rows = max(0, ws.max_row - start_row + 1)
        used_cols = max(0, ws.max_column - start_col + 1)
        rows = max(2, int(p["rows"])) if p.get("rows") else max(2, used_rows or 5)

        if explicit_headers:
            cols = max(len(explicit_headers), max(1, int(p["cols"])) if p.get("cols") else used_cols or len(explicit_headers))
        else:
            cols = max(1, int(p["cols"])) if p.get("cols") else max(1, used_cols or 3)

        seen_headers = set()
        for i in range(cols):
            header_cell = ws.cell(row=start_row, column=start_col + i)
            if i < len(explicit_headers):
                header_cell.value = explicit_headers[i]
            elif header_cell.value is None or str(header_cell.value).strip() == "":
                # Fall back to generic column name only when no explicit header supplied.
                header_cell.value = f"Column{i + 1}"
            header = str(header_cell.value or f"Column{i + 1}").strip() or f"Column{i + 1}"
            base_header = header
            suffix = 2
            while header.lower() in seen_headers:
                header = f"{base_header}_{suffix}"
                suffix += 1
            seen_headers.add(header.lower())
            header_cell.value = header

        end     = self._offset_cell(start, rows - 1, cols - 1)
        ref     = f"{start}:{end}"

        new_bounds = range_boundaries(ref)
        for existing in ws.tables.values():
            existing_ref = getattr(existing, "ref", "")
            if not existing_ref:
                continue
            if existing_ref.upper() == ref.upper():
                logger.info("Excel: table already exists for %s on %s; skipping duplicate.", ref, ws.title)
                return
            old_bounds = range_boundaries(existing_ref)
            if not (
                new_bounds[2] < old_bounds[0] or new_bounds[0] > old_bounds[2]
                or new_bounds[3] < old_bounds[1] or new_bounds[1] > old_bounds[3]
            ):
                raise ValueError(f"Cannot create overlapping Excel table range {ref}; existing table uses {existing_ref}.")

        existing_names = {
            name.lower()
            for sheet in self.wb.worksheets
            for name in sheet.tables.keys()
        }
        requested_name = str(p.get("table_name") or p.get("name") or "Table").strip()
        requested_name = re.sub(r"[^A-Za-z0-9_]", "_", requested_name)
        if not requested_name or not re.match(r"^[A-Za-z_]", requested_name):
            requested_name = "Table"
        idx = 1
        display_name = f"{requested_name}{idx}"
        while display_name.lower() in existing_names:
            idx += 1
            display_name = f"{requested_name}{idx}"

        tbl     = Table(displayName=display_name, ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showRowStripes=True
        )
        ws.add_table(tbl)
        self._set_current_sheet(ws)

    def _do_insert_chart(self, p):
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference
        chart_map = {
            "bar":  BarChart,
            "line": LineChart,
            "pie":  PieChart,
        }
        ChartClass = chart_map.get(p.get("chart_type", "bar"), BarChart)
        chart      = ChartClass()
        chart.title = "Chart"
        self.ws.add_chart(chart, p.get("start_cell", "E1"))

    def _do_create_pivot_table(self, p):
        logger.info(f"Pivot table from {p.get('source_range')} (requires win32com)")

    def _do_group_rows(self, p):
        self.ws.row_dimensions.group(int(p["start_row"]), int(p["end_row"]), hidden=False)

    def _do_ungroup_rows(self, p):
        for row_num in range(int(p["start_row"]), int(p["end_row"]) + 1):
            self.ws.row_dimensions[row_num].outlineLevel = 0

    def _do_insert_image(self, p):
        from openpyxl.drawing.image import Image
        import os
        path = p.get("path", "")
        if path and os.path.exists(path):
            img = Image(path)
            self.ws.add_image(img, p.get("cell", "A1"))

    # â”€â”€ View / Print â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    def _do_set_zoom(self, p):
        self.ws.sheet_view.zoomScale = int(p.get("level", 100))

    def _do_set_print_area(self, p):
        self.ws.print_area = p["range"]

    def _do_set_print_setup(self, p):
        self.ws.page_setup.orientation = p.get("orientation", "portrait")

    def _do_spell_check(self, p):
        logger.info("Spell check (requires win32com)")

    # â”€â”€ Save / Workbook â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€


    def _do_create_workbook(self, p):
        from openpyxl import Workbook
        self.wb = Workbook()
        self.ws = self.wb.active

    def _do_open_workbook(self, p):
        # Open/load is handled by server-level lifecycle.
        return

    def _do_save_workbook(self, p):
        # The office runner owns persistence and locked-file fallback. Treat
        # explicit save actions as lifecycle markers so they cannot fail early.
        return

    def _do_save_workbook_as(self, p):
        # The server resolves the final output path before execution and saves
        # once at the end. Saving here can corrupt the response path or fail
        # before the runner can apply its fallback for files open in Excel.
        return

    def _do_close_workbook(self, p):
        logger.info("Workbook closed (session ended)")

    # â”€â”€ Helpers â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    def _offset_cell(self, cell_ref, row_offset, col_offset):
        m = re.match(r'([A-Z]+)(\d+)', cell_ref.upper())
        if not m:
            return cell_ref
        col     = m.group(1)
        row     = int(m.group(2))
        col_num = sum(
            (ord(c) - 64) * (26 ** i)
            for i, c in enumerate(reversed(col))
        )
        new_col_num = col_num + col_offset
        new_col     = ""
        while new_col_num > 0:
            new_col     = chr((new_col_num - 1) % 26 + 65) + new_col
            new_col_num = (new_col_num - 1) // 26
        return f"{new_col}{row + row_offset}"

    def _set_table_ref(self, table, ref):
        table.ref = ref
        if getattr(table, "autoFilter", None) is not None:
            table.autoFilter.ref = ref

    def _shift_tables_for_insert_rows(self, row, amount):
        from openpyxl.utils import get_column_letter
        from openpyxl.utils.cell import range_boundaries

        for table in self.ws.tables.values():
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            if row <= min_row:
                min_row += amount
                max_row += amount
            elif min_row < row <= max_row:
                max_row += amount
            else:
                continue
            ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
            self._set_table_ref(table, ref)

    def _shift_tables_for_insert_cols(self, col, amount):
        from openpyxl.utils import get_column_letter
        from openpyxl.utils.cell import range_boundaries

        for table in self.ws.tables.values():
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            if col <= min_col:
                min_col += amount
                max_col += amount
            elif min_col < col <= max_col:
                max_col += amount
            else:
                continue
            ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
            self._set_table_ref(table, ref)
