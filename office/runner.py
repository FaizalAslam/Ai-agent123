import logging
import os
import traceback
from pathlib import Path

from office.constants import OFFICE_DEPENDENCIES, OFFICE_OUTPUTS
from office.intent import _action_names, _canonical_office_app
from office.file_resolver import _extract_named_file_path, _next_available_path
from executor.excel_executor import ExcelExecutor
from executor.word_executor import WordExecutor
from executor.ppt_executor import PowerPointExecutor


def _office_dependency_error(app_name):
    module_name, package_name = OFFICE_DEPENDENCIES.get(app_name, (None, None))
    if not module_name:
        return None
    try:
        __import__(module_name)
        return None
    except ModuleNotFoundError:
        return (
            f"{app_name.title()} support requires `{package_name}`. "
            f"Install it with `pip install {package_name}` or `pip install -r requirements.txt`."
        )


def _known_office_actions(app_name):
    cls = {
        "excel":      ExcelExecutor,
        "word":       WordExecutor,
        "powerpoint": PowerPointExecutor,
        "ppt":        PowerPointExecutor,
    }.get(app_name)
    if not cls:
        return set()
    return {name[4:] for name in dir(cls) if name.startswith("_do_")}


def _has_explicit_save_action(app_name, actions, command_text="", file_path=""):
    names = _action_names(actions)
    save_map = {
        "excel":      {"save_workbook",     "save_workbook_as"},
        "word":       {"save_document",     "save_document_as"},
        "powerpoint": {"save_presentation", "save_presentation_as"},
        "ppt":        {"save_presentation", "save_presentation_as"},
    }
    if names & save_map.get(app_name, set()):
        return True
    if (file_path or "").strip():
        return True
    if _extract_named_file_path(command_text, app_name):
        return True
    return False


def _run_office_actions(app_name, actions, file_path=None, command_text="", source_path=None):
    from modules import system_core

    app_name    = _canonical_office_app(app_name)
    output_path = str(Path((file_path or OFFICE_OUTPUTS.get(app_name, "output.xlsx"))).resolve())
    source_path = str(Path(source_path).resolve()) if source_path else ""
    executed    = []
    results     = []
    failures    = []
    opened      = False
    persisted   = False
    dependency_error = _office_dependency_error(app_name)

    logging.info(
        "Office execution start: app=%s source=%s output=%s actions=%s",
        app_name, source_path or "<new>", output_path, len(actions or []),
    )

    if dependency_error:
        failures.append(dependency_error)
        return {
            "success": False, "error_code": "OFFICE_DEPENDENCY_MISSING",
            "ok_count": 0, "total": len(actions or []),
            "executed": executed, "results": results, "failures": failures,
            "output_path": output_path, "opened": opened, "persisted": persisted,
            "dependency_error": dependency_error,
        }

    try:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        if app_name == "excel":
            from openpyxl import Workbook, load_workbook
            load_path = source_path or (output_path if os.path.exists(output_path) else "")
            wb = load_workbook(load_path) if load_path else Workbook()
            ws = wb.active
            setattr(wb, "_path", output_path)
            executor = ExcelExecutor(wb, ws)

            for idx, action in enumerate(actions or []):
                current_wb = getattr(executor, "wb", wb)
                setattr(current_wb, "_path", output_path)
                result = executor.run(action)
                current_wb = getattr(executor, "wb", current_wb)
                setattr(current_wb, "_path", output_path)
                action_name = action.get("action", "unknown")
                ok = isinstance(result, dict) and result.get("status") == "success"
                if ok:
                    executed.append(action_name)
                    results.append({
                        "action_index": idx, "action": action_name, "status": "success",
                        "target": action.get("range") or action.get("cell") or action.get("start_cell") or "",
                    })
                else:
                    msg = (result.get("message") or f"{action_name} failed") if isinstance(result, dict) else f"{action_name} failed"
                    err = (result.get("error_code") or "OFFICE_ACTION_FAILED") if isinstance(result, dict) else "OFFICE_ACTION_FAILED"
                    failures.append(msg)
                    results.append({
                        "action_index": idx, "action": action_name, "status": "failed",
                        "error_code": err, "message": msg,
                        "target": action.get("range") or action.get("cell") or action.get("start_cell") or "",
                    })

            final_obj   = getattr(executor, "wb", wb)
            save_method = final_obj.save

        elif app_name == "word":
            from docx import Document
            load_path = source_path or (output_path if os.path.exists(output_path) else "")
            doc = Document(load_path) if load_path else Document()
            setattr(doc, "_path", output_path)
            executor = WordExecutor(doc)

            for idx, action in enumerate(actions or []):
                current_doc = getattr(executor, "doc", doc)
                setattr(current_doc, "_path", output_path)
                result = executor.run(action)
                current_doc = getattr(executor, "doc", current_doc)
                setattr(current_doc, "_path", output_path)
                action_name = action.get("action", "unknown")
                ok = isinstance(result, dict) and result.get("status") == "success"
                if ok:
                    executed.append(action_name)
                    results.append({
                        "action_index": idx, "action": action_name, "status": "success",
                        "target": action.get("target") or action.get("text") or "",
                    })
                else:
                    msg = (result.get("message") or f"{action_name} failed") if isinstance(result, dict) else f"{action_name} failed"
                    err = (result.get("error_code") or "OFFICE_ACTION_FAILED") if isinstance(result, dict) else "OFFICE_ACTION_FAILED"
                    failures.append(msg)
                    results.append({
                        "action_index": idx, "action": action_name, "status": "failed",
                        "error_code": err, "message": msg,
                        "target": action.get("target") or action.get("text") or "",
                    })

            final_obj   = getattr(executor, "doc", doc)
            save_method = final_obj.save

        elif app_name == "powerpoint":
            from pptx import Presentation
            load_path = source_path or (output_path if os.path.exists(output_path) else "")
            prs = Presentation(load_path) if load_path else Presentation()
            setattr(prs, "_path", output_path)
            executor = PowerPointExecutor(prs)

            for idx, action in enumerate(actions or []):
                current_prs = getattr(executor, "prs", prs)
                setattr(current_prs, "_path", output_path)
                result = executor.run(action)
                current_prs = getattr(executor, "prs", current_prs)
                setattr(current_prs, "_path", output_path)
                action_name = action.get("action", "unknown")
                ok = isinstance(result, dict) and result.get("status") == "success"
                if ok:
                    executed.append(action_name)
                    results.append({
                        "action_index": idx, "action": action_name, "status": "success",
                        "target": action.get("slide_index") or action.get("target") or "",
                    })
                else:
                    msg = (result.get("message") or f"{action_name} failed") if isinstance(result, dict) else f"{action_name} failed"
                    err = (result.get("error_code") or "OFFICE_ACTION_FAILED") if isinstance(result, dict) else "OFFICE_ACTION_FAILED"
                    failures.append(msg)
                    results.append({
                        "action_index": idx, "action": action_name, "status": "failed",
                        "error_code": err, "message": msg,
                        "target": action.get("slide_index") or action.get("target") or "",
                    })

            final_obj   = getattr(executor, "prs", prs)
            save_method = final_obj.save

        else:
            return {
                "success": False, "error_code": "UNSUPPORTED_OFFICE_APP",
                "ok_count": len(executed), "total": len(actions or []),
                "executed": executed, "results": results,
                "failures": [f"Unsupported app: {app_name}"],
                "output_path": output_path, "persisted": False, "opened": False,
            }

        if not failures:
            try:
                save_method(output_path)
                persisted = Path(output_path).exists()
            except PermissionError:
                fallback_path = _next_available_path(output_path)
                save_method(fallback_path)
                output_path = fallback_path
                persisted   = Path(output_path).exists()
                logging.warning("%s target was locked. Saved to fallback path: %s", app_name, output_path)
            except Exception as exc:
                failures.append(str(exc))
                logging.error("Office save failed for %s: %s", output_path, exc)

        if not failures and not persisted:
            failures.append("Save did not create a file on disk.")

    except Exception as exc:
        failures.append(str(exc))
        logging.error("Office execution failed: %s\n%s", exc, traceback.format_exc())

    if not failures and persisted and os.path.exists(output_path):
        try:
            opened = bool(system_core.open_path(output_path))
        except Exception as exc:
            logging.warning("Could not open output file %s: %s", output_path, exc)
            opened = False

    success = not failures and persisted and Path(output_path).exists()
    logging.info(
        "Office execution result: app=%s success=%s saved=%s output=%s",
        app_name, success, persisted, output_path,
    )

    return {
        "success":    success,
        "error_code": "" if success else (
            "OFFICE_ACTION_FAILED"
            if results and any(r.get("status") in {"fail", "failed"} for r in results)
            else "OFFICE_SAVE_FAILED"
        ),
        "ok_count":  len(executed),
        "total":     len(actions or []),
        "executed":  executed,
        "results":   results,
        "failures":  failures,
        "output_path": output_path,
        "persisted": persisted,
        "opened":    opened,
    }
